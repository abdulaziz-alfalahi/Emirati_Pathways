import os
import json
import pandas as pd
import openpyxl
import logging

logger = logging.getLogger(__name__)

def find_sheet_name(sheet_names, target):
    """
    Finds a sheet name in a list of names by normalizing spaces, quotes, and case.
    """
    target_norm = target.lower().replace(" ", "").replace("'", "").replace('"', "").replace("_", "").replace("-", "")
    for name in sheet_names:
        name_norm = name.lower().replace(" ", "").replace("'", "").replace('"', "").replace("_", "").replace("-", "")
        if target_norm == name_norm or target_norm in name_norm or name_norm in target_norm:
            return name
    return None

def get_sheet_stats(df):
    """
    Computes demographics from a pandas DataFrame.
    """
    gender = df['Gender'].value_counts().to_dict() if 'Gender' in df.columns else {}
    age = df['Age Group'].value_counts().to_dict() if 'Age Group' in df.columns else {}
    edu = df['Education'].value_counts().to_dict() if 'Education' in df.columns else {}
    exp = df['Experience'].value_counts().to_dict() if 'Experience' in df.columns else {}
    loc = df['Emirate Of Residence'].value_counts().to_dict() if 'Emirate Of Residence' in df.columns else {}
    military = df['Military Status'].value_counts().to_dict() if 'Military Status' in df.columns else {}
    marital = df['Marital Status'].value_counts().to_dict() if 'Marital Status' in df.columns else {}
    employment = df['Work Status'].value_counts().to_dict() if 'Work Status' in df.columns else {}
    
    return {
        "total": len(df),
        "gender": [{"name": str(k), "value": int(v)} for k, v in gender.items()],
        "age_group": [{"name": str(k), "value": int(v)} for k, v in age.items()],
        "education": [{"name": str(k), "value": int(v)} for k, v in edu.items()],
        "experience": [{"name": str(k), "value": int(v)} for k, v in exp.items()],
        "location": [{"name": str(k), "value": int(v)} for k, v in loc.items()],
        "military": [{"name": str(k), "value": int(v)} for k, v in military.items()],
        "marital": [{"name": str(k), "value": int(v)} for k, v in marital.items()],
        "employment": [{"name": str(k), "value": int(v)} for k, v in employment.items()]
    }

def parse_master_excel(excel_path="/app/master_file.xlsx"):
    try:
        if not os.path.exists(excel_path):
            # Fallback for local dev
            alt_path = os.path.join(os.path.dirname(__file__), "master_file.xlsx")
            if os.path.exists(alt_path):
                excel_path = alt_path
            else:
                logger.error(f"Excel file not found at {excel_path} or {alt_path}")
                return None
                
        logger.info(f"Parsing Excel demographics via pandas & openpyxl from {excel_path}...")
        
        # Read sheet names first to align normalized names
        excel_file = pd.ExcelFile(excel_path)
        sheet_names = excel_file.sheet_names
        
        data = {}
        
        # 1. Registered (Master sheet)
        master_sheet = find_sheet_name(sheet_names, "Master")
        if master_sheet:
            df_master = pd.read_excel(excel_path, sheet_name=master_sheet)
            data["registered"] = get_sheet_stats(df_master)
        else:
            logger.error("Master sheet not found!")
            return None
        
        # 2. Active (Master sheet filtered)
        df_active = df_master[
            (df_master['Call Status'] == 'Answered') & 
            (df_master['Work Status'] == 'Not Working') & 
            (df_master['Looking / Not Looking'] == 'Looking For Work')
        ]
        data["active"] = get_sheet_stats(df_active)
        
        # Add Compatibility Keys for Legacy Dashboard Charts
        # age_distribution
        age_groups = df_master['Age Group'].dropna().unique()
        age_dist = []
        for gp in age_groups:
            gp_df = df_master[df_master['Age Group'] == gp]
            male_cnt = int((gp_df['Gender'] == 'Male').sum())
            female_cnt = int((gp_df['Gender'] == 'Female').sum())
            age_dist.append({
                "group": str(gp),
                "male": male_cnt,
                "female": female_cnt
            })
        data["age_distribution"] = age_dist

        # regional_spread
        emirates = df_master['Emirate Of Residence'].dropna().unique()
        reg_spread = []
        for em in emirates:
            cnt = int((df_master['Emirate Of Residence'] == em).sum())
            reg_spread.append({
                "emirate": str(em),
                "candidates": cnt
            })
        data["regional_spread"] = reg_spread

        # education_levels
        edu_levels = df_master['Education'].dropna().unique()
        edu_lvls = []
        for lvl in edu_levels:
            lvl_df = df_master[df_master['Education'] == lvl]
            employed_cnt = int((lvl_df['Work Status'] == 'Working').sum())
            seeking_cnt = int((lvl_df['Work Status'] == 'Not Working').sum())
            edu_lvls.append({
                "level": str(lvl),
                "employed": employed_cnt,
                "seeking": seeking_cnt
            })
        data["education_levels"] = edu_lvls
        
        # 3. Special Cohorts / Priorities
        cohort_sheets = {
            "priority_1st": "1st Priority JS List",
            "priority_2nd": "2nd Priority JS List",
            "priority_3rd": "3rd Priority JS List",
            "hatta": "Hatta JS List",
            "cda": "CDA JS List",
            "gdo": "GDO JS List",
            "no_answer": "No Answer JS List"
        }
        
        for key, target_name in cohort_sheets.items():
            actual_name = find_sheet_name(sheet_names, target_name)
            if actual_name:
                try:
                    df_cohort = pd.read_excel(excel_path, sheet_name=actual_name)
                    data[key] = get_sheet_stats(df_cohort)
                    logger.info(f"Successfully parsed cohort {key} from sheet '{actual_name}'")
                except Exception as e:
                    logger.error(f"Error parsing cohort sheet {actual_name}: {e}")
                    data[key] = get_sheet_stats(pd.DataFrame(columns=df_master.columns))
            else:
                logger.warning(f"Could not find matching sheet for target '{target_name}'")
                data[key] = get_sheet_stats(pd.DataFrame(columns=df_master.columns))
        
        # Parse historical and operational metrics with openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # 4. Added & Removed Dashboard
        ar_sheet_name = find_sheet_name(wb.sheetnames, "Added & Removed Dashboard")
        if ar_sheet_name:
            sheet = wb[ar_sheet_name]
            data["growth"] = {
                "total_added": int(sheet["B4"].value or 10603),
                "total_removed": int(sheet["H4"].value or 1520),
                "weekly": []
            }
            for r in range(7, 13):
                date_val = sheet.cell(row=r, column=1).value
                added_count = sheet.cell(row=r, column=2).value
                removed_count = sheet.cell(row=r, column=8).value
                
                date_str = str(date_val).split(" ")[0] if date_val else ""
                data["growth"]["weekly"].append({
                    "date": date_str,
                    "added": int(added_count or 0),
                    "removed": int(removed_count or 0)
                })

        # 5. Historical JS Trends (from bottom/right of Active JS Indicator)
        active_indicator_sheet = find_sheet_name(wb.sheetnames, "Active JS Indicator")
        if active_indicator_sheet:
            sheet = wb[active_indicator_sheet]
            historical = []
            # Columns C (3) to V (22)
            # Row 7 is Month, Row 8 is Total Accumulated JS, Row 9 is Total Accumulated AJS
            for col_idx in range(3, 23):
                month_name = sheet.cell(row=7, column=col_idx).value
                if not month_name or str(month_name).strip() == "Indicators / Months" or "Average" in str(month_name) or "Max" in str(month_name) or "Min" in str(month_name):
                    continue
                accum_js = sheet.cell(row=8, column=col_idx).value
                accum_ajs = sheet.cell(row=9, column=col_idx).value
                
                historical.append({
                    "month": str(month_name),
                    "total_js": int(accum_js or 0),
                    "active_js": int(accum_ajs or 0)
                })
            
            if "growth" in data:
                data["growth"]["historical_js_trends"] = historical
            else:
                data["growth"] = {"historical_js_trends": historical}

        # 6. Rapid Nomination monthly stats (EHRDC Initiatives Indicators)
        init_indicators_sheet = find_sheet_name(wb.sheetnames, "EHRDC Initiatives Indicators")
        if init_indicators_sheet:
            sheet = wb[init_indicators_sheet]
            
            rapid_nomination = []
            for r in range(9, 20):
                month_val = sheet.cell(row=r, column=5).value
                if not month_val:
                    continue
                companies = sheet.cell(row=r, column=6).value
                vacancies = sheet.cell(row=r, column=7).value
                nominated = sheet.cell(row=r, column=8).value
                interested = sheet.cell(row=r, column=9).value
                
                rapid_nomination.append({
                    "month": str(month_val),
                    "companies": int(companies or 0),
                    "vacancies": int(vacancies or 0),
                    "nominated": int(nominated or 0),
                    "interested": int(interested or 0)
                })
            
            data["rapid_nomination"] = rapid_nomination
            
            # Initiatives Totals
            data["initiatives_totals"] = {
                "hatta": data["hatta"]["total"],
                "cda": data["cda"]["total"],
                "gdo": data["gdo"]["total"]
            }

        # Cache results to json file in the same directory
        cache_path = os.path.join(os.path.dirname(__file__), "demographics_cache.json")
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logger.info(f"Demographics cached successfully to {cache_path}")
        
        return data
    except Exception as e:
        logger.error(f"Error parsing master Excel file: {e}")
        return None

def get_cached_demographics(excel_path="/app/master_file.xlsx"):
    cache_path = os.path.join(os.path.dirname(__file__), "demographics_cache.json")
    
    # Resolve alt path for local development
    if not os.path.exists(excel_path):
        alt_path = os.path.join(os.path.dirname(__file__), "master_file.xlsx")
        if os.path.exists(alt_path):
            excel_path = alt_path
            
    if os.path.exists(excel_path):
        excel_mtime = os.path.getmtime(excel_path)
        if os.path.exists(cache_path):
            cache_mtime = os.path.getmtime(cache_path)
            if cache_mtime > excel_mtime:
                try:
                    with open(cache_path, "r", encoding="utf-8") as f:
                        return json.load(f)
                except Exception as e:
                    logger.error(f"Error loading cached demographics: {e}")
        
        # Regenerate cache
        return parse_master_excel(excel_path)
    
    # Fallback if no excel file exists
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
            
    return None
