#!/usr/bin/env python3
"""Import a CRM "Main Master File" workbook into the platform CRM.

Usage:  python backend/scripts/import_crm_master.py <Main_Master_File.xlsx>

What it does (idempotent — safe to re-run with the same or a newer file):
  1. Master sheet  -> upsert users (id = real Emirates ID) + candidate_profiles
     CRM fields (call/work/looking status, CV status, remarks, segments, ...).
     Existing users keep their platform data; contact fields fill only if NULL.
  2. Segment list sheets -> candidate_profiles.crm_segments (jsonb array).
  3. "Add & Remove Pivot" -> crm_roster_history (weekly + monthly series).
  4. Profiles previously on the CRM roster (crm_reference set) but absent from
     this file get crm_segments = [] (off-roster; nothing is deleted).

Requires migration 044 (RAN live 2026-08-03).
"""
import json
import os
import sys
from datetime import datetime, date

import psycopg2
from dotenv import dotenv_values
from openpyxl import load_workbook

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
ENV = dotenv_values(os.path.join(REPO, 'backend', '.env'))

SEGMENT_SHEETS = {
    'Active JS List': 'active',
    '1st Priority JS List': 'priority_1',
    '2nd Priority JS List': 'priority_2',
    '3rd Priority JS List': 'priority_3',
    'Hatta JS List': 'hatta',
    'CDA JS List': 'cda',
    'Special Request JS List': 'special_request',
    'GDO JS List': 'gdo',          # sheet name may carry a trailing space
    'No Answer JS List': 'no_answer',
    'Prev Employed 21-24 List': 'prev_employed_21_24',
    'Never Employed 21-24 List': 'never_employed_21_24',
}

MONTHS = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6, 'june': 6,
          'jul': 7, 'july': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}


def _s(v, maxlen=None):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() in ('none', 'nan', 'not defined', 'not specified &'):
        return None
    return s[:maxlen] if maxlen else s


def _eid(v):
    s = _s(v)
    if not s:
        return None
    s = s.split('.')[0].strip()
    return s if s.isdigit() and len(s) == 15 else None


def _dt(v):
    if isinstance(v, datetime):
        return v
    s = _s(v)
    if not s:
        return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d-%b-%Y', '%m/%d/%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(s.split('.')[0], fmt)
        except ValueError:
            continue
    return None


def _bool(v):
    s = _s(v)
    if s is None:
        return None
    return s.lower() in ('true', 'yes', '1')


def _month_label_to_date(label):
    """"May '24" / "May'25" / "July'25" -> date(yyyy, m, 1)."""
    s = label.replace("'", ' ').replace('  ', ' ').strip()
    parts = s.split()
    if len(parts) != 2:
        return None
    m = MONTHS.get(parts[0].strip().lower()[:4].rstrip('.')) or MONTHS.get(parts[0].strip().lower()[:3])
    try:
        y = 2000 + int(parts[1])
    except ValueError:
        return None
    return date(y, m, 1) if m else None


def parse_sheet_rows(ws):
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h else '' for h in next(rows)]
    idx = {h: i for i, h in enumerate(header)}  # first occurrence wins

    out = {}
    for r in rows:
        eid = _eid(r[idx['EID']])
        if not eid:
            continue
        g = lambda col: r[idx[col]] if col in idx else None
        out[eid] = {
            'crm_reference': _s(g('Reference'), 30),
            'full_name': _s(g('Full Name')),
            'gender': _s(g('Gender'), 12),
            'dob': _dt(g('Date Of Birth')),
            'age_group': _s(g('Age Group'), 20),
            'education_level': _s(g('Education'), 80),
            'crm_registered_on': _dt(g('Registered On')),
            'job_seeker_type': _s(g('Job Seeker Type'), 50),
            'job_seeker_date': _dt(g('Job Seeker Date')),
            'is_student': _bool(g('Is Student')),
            'emirate_of_residence': _s(g('Emirate Of Residence'), 60),
            'specialization': _s(g('Specialization'), 160),
            'phone': _s(g('Ph No'), 30),
            'email': _s(g('Email')),
            'date_of_call': _dt(g('Date Of Call')),
            'call_status': _s(g('Call Status'), 50),
            'work_status': _s(g('Work Status'), 50),
            'looking_status': _s(g('Looking / Not Looking'), 50),
            'salary_expectations': _s(g('Salary Expectations'), 60),
            'cv_status': _s(g('CV Status'), 80),
            'remarks': _s(g('Remarks')),
            'candidates_source': _s(g('Candidates’ Source') or g("Candidates' Source"), 80),
        }
    return out


def parse_master(wb):
    return parse_sheet_rows(wb['Master'])


def parse_segments(wb, master):
    """Collect segment membership. Some segment sheets (CDA / Special Request /
    GDO) carry candidates who are NOT in the Master sheet — parse their full
    rows too and merge them into the roster (Master data wins on conflict)."""
    seg_by_eid = {}
    names = {n.strip(): n for n in wb.sheetnames}
    for wanted, seg in SEGMENT_SHEETS.items():
        actual = names.get(wanted)
        if not actual:
            print(f'  ! segment sheet missing: {wanted}')
            continue
        extra = parse_sheet_rows(wb[actual])
        n = 0
        for eid, row in extra.items():
            seg_by_eid.setdefault(eid, []).append(seg)
            n += 1
            if eid not in master:
                master[eid] = row
        print(f'  segment {seg}: {n}')
    return seg_by_eid


def parse_roster_history(wb):
    ws = wb['Add & Remove Pivot']
    months, weeks = {}, {}
    for r in ws.iter_rows(values_only=True):
        # Months block: B..E = Date, Added, Removed, Total
        d, a, rem = r[1] if len(r) > 1 else None, r[2] if len(r) > 2 else None, r[3] if len(r) > 3 else None
        if d is not None and _s(a) and str(a).strip().isdigit():
            if isinstance(d, datetime):
                key = date(d.year, d.month, 1)
                if key not in months:
                    months[key] = [0, 0, f"{d.strftime('%b')} '{str(d.year)[2:]}"]
                months[key][0] += int(a); months[key][1] += int(rem or 0)
            else:
                label = _s(d)
                if label and label.lower() != 'total':
                    md = _month_label_to_date(label)
                    if md:
                        months.setdefault(md, [0, 0, label])
                        months[md][0] += int(a); months[md][1] += int(rem or 0)
        # Weeks block: G..J = Date, Removed, Added, Total  (swapped vs months!)
        wd = r[6] if len(r) > 6 else None
        wrem, wadd = r[7] if len(r) > 7 else None, r[8] if len(r) > 8 else None
        if wd is not None and _s(wadd) and str(wadd).strip().isdigit():
            if isinstance(wd, datetime):
                key = wd.date()
            else:
                label = _s(wd)
                if not label or label.lower() == 'total':
                    continue
                try:  # "19-May" style, first weeks of the series -> 2024
                    key = datetime.strptime(f'{label}-2024', '%d-%b-%Y').date()
                except ValueError:
                    continue
            if key in weeks:  # the workbook has duplicate week labels — sum them
                weeks[key][0] += int(wadd); weeks[key][1] += int(wrem or 0)
            else:
                weeks[key] = [int(wadd), int(wrem or 0), key.strftime('%d %b %y')]
    return months, weeks


def main(path):
    print(f'Loading {path} ...')
    wb = load_workbook(path, read_only=True, data_only=True)
    master = parse_master(wb)
    print(f'Master rows: {len(master)}')
    segs = parse_segments(wb, master)
    print(f'Roster incl. segment-only sheets: {len(master)}')
    months, weeks = parse_roster_history(wb)
    print(f'History: {len(months)} months, {len(weeks)} weeks')

    conn = psycopg2.connect(host=ENV['DB_HOST'], port=ENV['DB_PORT'], dbname=ENV['DB_NAME'],
                            user=ENV['DB_USER'], password=ENV['DB_PASSWORD'])
    cur = conn.cursor()

    cur.execute("SELECT emirates_id_enc, id FROM users WHERE emirates_id_enc IS NOT NULL")
    existing = {r[0].strip(): r[1] for r in cur.fetchall() if r[0]}

    created = updated = 0
    for eid, m in master.items():
        seg_json = json.dumps(segs.get(eid, []))
        if eid not in existing:
            # New platform user: real EID as id — UAE Pass will bind to it later.
            cur.execute("""
                INSERT INTO users (id, emirates_id_enc, role, user_type, full_name,
                                   email, phone, is_active, created_at, updated_at)
                VALUES (%s, %s, 'candidate', 'candidate', %s, %s, %s, true, NOW(), NOW())
                ON CONFLICT (id) DO NOTHING
            """, (eid, eid, m['full_name'], m['email'], m['phone']))
            created += 1
        else:
            cur.execute("""
                UPDATE users SET
                    full_name = COALESCE(full_name, %s),
                    email = COALESCE(email, %s),
                    phone = COALESCE(phone, %s),
                    updated_at = NOW()
                WHERE emirates_id_enc = %s
            """, (m['full_name'], m['email'], m['phone'], eid))

        cur.execute("SELECT 1 FROM candidate_profiles WHERE user_id = %s", (eid,))
        if cur.fetchone():
            cur.execute("""
                UPDATE candidate_profiles SET
                    crm_reference = %(crm_reference)s,
                    crm_segments = %(crm_segments)s::jsonb,
                    call_status = %(call_status)s,
                    work_status = %(work_status)s,
                    job_seeker_type = %(job_seeker_type)s,
                    looking_status = %(looking_status)s,
                    cv_status = %(cv_status)s,
                    date_of_call = %(date_of_call)s,
                    candidates_source = %(candidates_source)s,
                    education_level = %(education_level)s,
                    specialization = %(specialization)s,
                    emirate_of_residence = %(emirate_of_residence)s,
                    age_group = %(age_group)s,
                    gender = %(gender)s,
                    is_student = %(is_student)s,
                    salary_expectations = %(salary_expectations)s,
                    job_seeker_date = %(job_seeker_date)s,
                    crm_registered_on = %(crm_registered_on)s,
                    counseling_remarks = COALESCE(%(remarks)s, counseling_remarks),
                    dob = COALESCE(dob, %(dob)s),
                    phone = COALESCE(phone, %(phone)s),
                    updated_at = NOW()
                WHERE user_id = %(eid)s
            """, {**m, 'eid': eid, 'crm_segments': seg_json})
        else:
            cur.execute("""
                INSERT INTO candidate_profiles (user_id, full_name, crm_reference, crm_segments,
                    call_status, work_status, job_seeker_type, looking_status, cv_status,
                    date_of_call, candidates_source, education_level, specialization,
                    emirate_of_residence, age_group, gender, is_student, salary_expectations,
                    job_seeker_date, crm_registered_on, counseling_remarks, dob, phone,
                    created_at, updated_at)
                VALUES (%(eid)s, %(full_name)s, %(crm_reference)s, %(crm_segments)s::jsonb,
                    %(call_status)s, %(work_status)s, %(job_seeker_type)s, %(looking_status)s,
                    %(cv_status)s, %(date_of_call)s, %(candidates_source)s, %(education_level)s,
                    %(specialization)s, %(emirate_of_residence)s, %(age_group)s, %(gender)s,
                    %(is_student)s, %(salary_expectations)s, %(job_seeker_date)s,
                    %(crm_registered_on)s, %(remarks)s, %(dob)s, %(phone)s, NOW(), NOW())
            """, {**m, 'eid': eid, 'crm_segments': seg_json})
        updated += 1
        if updated % 500 == 0:
            conn.commit()
            print(f'  ... {updated}')

    # Off-roster: had a CRM reference before, absent from this file.
    cur.execute("""
        UPDATE candidate_profiles SET crm_segments = '[]'::jsonb, updated_at = NOW()
        WHERE crm_reference IS NOT NULL
          AND crm_segments <> '[]'::jsonb
          AND user_id NOT IN %s
    """, (tuple(master.keys()),))
    off_roster = cur.rowcount

    for d, (a, rem, label) in months.items():
        cur.execute("""
            INSERT INTO crm_roster_history (period_type, period_date, period_label, added, removed, total, source)
            VALUES ('month', %s, %s, %s, %s, %s, %s)
            ON CONFLICT (period_type, period_date)
            DO UPDATE SET added = EXCLUDED.added, removed = EXCLUDED.removed,
                          total = EXCLUDED.total, source = EXCLUDED.source
        """, (d, label, a, rem, a + rem, os.path.basename(path)))
    for d, (a, rem, label) in weeks.items():
        cur.execute("""
            INSERT INTO crm_roster_history (period_type, period_date, period_label, added, removed, total, source)
            VALUES ('week', %s, %s, %s, %s, %s, %s)
            ON CONFLICT (period_type, period_date)
            DO UPDATE SET added = EXCLUDED.added, removed = EXCLUDED.removed,
                          total = EXCLUDED.total, source = EXCLUDED.source
        """, (d, label, a, rem, a + rem, os.path.basename(path)))

    conn.commit()
    print(f'DONE: {created} users created, {updated} profiles upserted, '
          f'{off_roster} marked off-roster, {len(months)}+{len(weeks)} history rows')
    conn.close()


if __name__ == '__main__':
    if len(sys.argv) != 2:
        sys.exit('usage: import_crm_master.py <Main_Master_File.xlsx>')
    main(sys.argv[1])
